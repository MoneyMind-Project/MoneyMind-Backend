from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from rest_framework import status
from moneymind_apps.balances.models import *
from moneymind_apps.alerts.models import Alert, AlertType
from moneymind_apps.users.models import *
from django.utils import timezone
from moneymind_apps.reports.models import *
from moneymind_apps.movements.utils.services.gemini_api import *
from moneymind_apps.movements.models import *
from moneymind_apps.alerts.views import get_recurring_payment_reminders
import math




from rest_framework import status
from django.http import HttpResponse
from django.db.models import Sum, Count, Q
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
import io

# Para Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Para PDF
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT


class UnifiedDashboardAnalyticsView(APIView):
    permission_classes = [AllowAny]
    """
    Endpoint unificado que devuelve todos los datos de analytics del dashboard
    en una sola respuesta, optimizando las llamadas al backend.
    """

    def get(self, request):
        user_id = request.query_params.get('user_id')
        year = request.query_params.get('year', None)
        month = request.query_params.get('month', None)

        if not user_id:
            return Response(
                {"error": "user_id es requerido"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Usar año actual si no se proporciona
            if not year:
                year = datetime.now().year
            else:
                year = int(year)

            # Usar mes actual si no se proporciona
            if not month:
                month = datetime.now().month
            else:
                month = int(month)
                if month < 1 or month > 12:
                    raise ValueError("Mes inválido")

        except ValueError as e:
            return Response(
                {"error": f"Parámetros inválidos: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        current_date = datetime.now()
        current_month = current_date.month if current_date.year == year else 12

        # Construir respuesta unificada
        analytics_data = {
            "success": True,
            "user_id": user_id,
            "year": year,
            "month": month,
            "current_month": current_month,
            "data": {
                "monthly_predictions": self._get_monthly_predictions(user_id, year, current_month),
                "expenses_by_category": self._get_expenses_by_category(user_id, month, year),
                "expenses_by_parent_category": self._get_expenses_by_parent_category(user_id, month, year),
                "savings_evolution": self._get_savings_evolution(user_id, year),
                "essential_vs_non_essential": self._get_essential_vs_non_essential(user_id, year)
            }
        }

        return Response(analytics_data, status=status.HTTP_200_OK)

    def _get_monthly_predictions(self, user_id, year, current_month):
        """
        Predicciones de gastos mensuales
        """
        monthly_expenses = []

        for month in range(1, 13):
            month_total = Expense.objects.filter(
                user_id=user_id,
                date__month=month,
                date__year=year
            ).aggregate(total=Sum('total'))['total']

            monthly_expenses.append({
                "month": month,
                "real": float(month_total) if month_total else None
            })

        # Calcular predicción
        months_with_data = [m for m in monthly_expenses if m["real"] is not None and m["real"] > 0]

        if len(months_with_data) >= 3:
            last_3_months = months_with_data[-3:]
            avg_expense = sum(m["real"] for m in last_3_months) / 3
        elif len(months_with_data) > 0:
            avg_expense = sum(m["real"] for m in months_with_data) / len(months_with_data)
        else:
            avg_expense = 0

        current_month_value = monthly_expenses[current_month - 1]["real"] if current_month <= 12 else None

        predictions = []
        for i, month_data in enumerate(monthly_expenses):
            month = month_data["month"]

            if month < current_month:
                predictions.append(None)
            elif month == current_month:
                predictions.append(current_month_value)
            else:
                months_ahead = month - current_month
                base_value = current_month_value if current_month_value else avg_expense
                predicted_value = base_value * (1.02 ** months_ahead)
                predictions.append(round(predicted_value, 2))

        result = []
        for i, month_data in enumerate(monthly_expenses):
            result.append({
                "month": month_data["month"],
                "real": month_data["real"],
                "prediction": predictions[i]
            })

        return result

    def _get_expenses_by_category(self, user_id, month, year):
        """
        Gastos por categoría (16 categorías)
        """
        expenses = Expense.objects.filter(
            user_id=user_id,
            date__month=month,
            date__year=year
        )

        category_totals = expenses.values('category').annotate(
            total=Sum('total')
        )

        all_categories = {cat.value: 0 for cat in Category}

        for item in category_totals:
            all_categories[item['category']] = float(item['total'])

        result = [
            {
                "category": category,
                "total": total
            }
            for category, total in all_categories.items()
        ]

        return result

    def _get_expenses_by_parent_category(self, user_id, month, year):
        """
        Gastos por categoría padre (4 grupos principales)
        """
        expenses = Expense.objects.filter(
            user_id=user_id,
            date__month=month,
            date__year=year
        )

        parent_totals = {parent.value: 0 for parent in CategoryParent}

        for expense in expenses:
            try:
                category = Category(expense.category)
                parent_category = CATEGORY_PARENT_MAP.get(category)

                if parent_category:
                    parent_totals[parent_category.value] += float(expense.total)

            except ValueError:
                continue

        result = [
            {
                "parent_category": parent_category,
                "total": total
            }
            for parent_category, total in parent_totals.items()
        ]

        return result

    def _get_savings_evolution(self, user_id, year):
        """
        Evolución del ahorro mes a mes
        """
        history = UserBalanceHistory.objects.filter(
            user_id=user_id,
            date__year=year
        ).order_by('date')

        monthly_savings = []
        previous_amount = None

        for record in history:
            if previous_amount is not None:
                saving = float(record.amount) - previous_amount
            else:
                saving = 0

            monthly_savings.append({
                "month": record.date.month,
                "date": record.date.isoformat(),
                "balance": float(record.amount),
                "saving": saving
            })

            previous_amount = float(record.amount)

        return monthly_savings

    def _get_essential_vs_non_essential(self, user_id, year):
        """
        Gastos esenciales vs no esenciales por mes
        """
        monthly_data = []

        for month in range(1, 13):
            expenses = Expense.objects.filter(
                user_id=user_id,
                date__month=month,
                date__year=year
            )

            esencial_total = 0
            no_esencial_total = 0

            for expense in expenses:
                try:
                    category = Category(expense.category)
                    expense_type = CATEGORY_EXPENSE_TYPE_MAP.get(category)

                    if expense_type == ExpenseType.ESENCIAL:
                        esencial_total += float(expense.total)
                    elif expense_type == ExpenseType.NO_ESENCIAL:
                        no_esencial_total += float(expense.total)

                except ValueError:
                    continue

            monthly_data.append({
                "month": month,
                "esencial": esencial_total,
                "no_esencial": no_esencial_total
            })

        return monthly_data


class DashboardOverviewView(APIView):
    permission_classes = [AllowAny]
    """
    Retorna los KPIs principales del dashboard
    """

    def get(self, request):
        user_id = request.query_params.get('user_id')
        month = request.query_params.get('month')
        year = request.query_params.get('year')

        if not user_id:
            return Response(
                {"error": "user_id es requerido"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            if not month:
                month = datetime.now().month
            else:
                month = int(month)

            if not year:
                year = datetime.now().year
            else:
                year = int(year)

        except ValueError as e:
            return Response(
                {"error": f"Parámetros inválidos: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # 1. Total gastado este mes
        total_gastado = Expense.objects.filter(
            user_id=user_id,
            date__month=month,
            date__year=year
        ).aggregate(total=Sum('total'))['total'] or Decimal('0')

        # 2. Categoría más alta
        categoria_alta = Expense.objects.filter(
            user_id=user_id,
            date__month=month,
            date__year=year
        ).values('category').annotate(
            total=Sum('total')
        ).order_by('-total').first()

        categoria_mas_alta = None
        if categoria_alta:
            from moneymind_apps.movements.models import CATEGORY_LABELS
            try:
                cat_enum = Category(categoria_alta['category'])
                categoria_mas_alta = {
                    "category": categoria_alta['category'],
                    "label": CATEGORY_LABELS.get(cat_enum, categoria_alta['category']),
                    "total": float(categoria_alta['total'])
                }
            except ValueError:
                categoria_mas_alta = None

        # 3. Presupuesto restante (balance actual)
        try:
            balance = Balance.objects.get(user_id=user_id)
            presupuesto_restante = float(balance.current_amount)
        except Balance.DoesNotExist:
            presupuesto_restante = 0

        # 👈 NUEVO: Calcular presupuesto inicial del mes
        # Presupuesto inicial = balance actual + total gastado este mes
        presupuesto_inicial_mes = presupuesto_restante + float(total_gastado)

        # 4. Proyección del próximo mes (promedio de últimos 3 meses)
        last_3_months_totals = []

        for i in range(1, 4):
            target_month = month - i
            target_year = year

            if target_month <= 0:
                target_month += 12
                target_year -= 1

            month_total = Expense.objects.filter(
                user_id=user_id,
                date__month=target_month,
                date__year=target_year
            ).aggregate(total=Sum('total'))['total'] or Decimal('0')

            last_3_months_totals.append(float(month_total))

        # 🔹 Lógica mejorada de proyección
        valid_months = [v for v in last_3_months_totals if v > 0]

        if len(valid_months) >= 2:
            # Promedio de los meses que sí tienen datos
            proyeccion = sum(valid_months) / len(valid_months)
        elif len(valid_months) == 1:
            # Si solo hay un mes, usa ese valor (lo repetimos virtualmente)
            proyeccion = valid_months[0]
        else:
            # Si no hay ningún dato previo, usa el gasto actual como referencia
            proyeccion = float(total_gastado)

        # 🔸 Redondear hacia arriba si tiene decimales
        proyeccion = math.ceil(proyeccion)

        # Verificar y crear alerta si es necesario
        # 👈 CAMBIO: Pasar presupuesto_inicial_mes en lugar de presupuesto_restante
        self.check_budget_alert(user_id, month, year, float(total_gastado), presupuesto_inicial_mes)

        return Response(
            {
                "success": True,
                "data": {
                    "total_gastado_mes": float(total_gastado),
                    "categoria_mas_alta": categoria_mas_alta,
                    "presupuesto_restante": presupuesto_restante,
                    "proyeccion_proximo_mes": round(proyeccion, 2)
                },
                "month": month,
                "year": year
            },
            status=status.HTTP_200_OK
        )

    def check_budget_alert(self, user_id, month, year, total_gastado, presupuesto_inicial_mes):
        """
        Verifica si se debe crear una alerta de presupuesto
        """
        # Calcular si se gastó más de 2/3 del presupuesto inicial del mes
        if presupuesto_inicial_mes > 0:
            umbral = presupuesto_inicial_mes * (2 / 3)

            if total_gastado >= umbral:
                # Verificar si ya existe una alerta para este mes
                alert_exists = Alert.objects.filter(
                    user_id=user_id,
                    alert_type=AlertType.RISK.value,
                    target_month=month,
                    target_year=year
                ).exists()

                if not alert_exists:
                    # Calcular cuánto queda del presupuesto inicial
                    restante = presupuesto_inicial_mes - total_gastado
                    porcentaje = (total_gastado / presupuesto_inicial_mes) * 100

                    Alert.objects.create(
                        user_id=user_id,
                        alert_type=AlertType.RISK.value,
                        message=f"Has gastado 2/3 de tu presupuesto inicial este mes (S/ {presupuesto_inicial_mes:.2f}).",
                        target_month=month,
                        target_year=year,
                        seen=False
                    )

class HomeDashboardView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        user_id = request.query_params.get('user_id')

        if not user_id:
            return Response(
                {"error": "user_id es requerido"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response(
                {"error": "Usuario no encontrado"},
                status=status.HTTP_404_NOT_FOUND
            )

        # Obtener mes y año actuales
        now = timezone.now()
        current_month = now.month
        current_year = now.year

        # 2. Obtener tip semanal (temporal - datos inventados)
        weekly_tip = self._get_weekly_tip(int(user_id))

        # 3. Obtener datos de presupuesto
        budget_data = self._get_budget_data(user_id, current_month, current_year)

        # 4. Obtener gastos diarios del mes
        daily_expenses = self._get_daily_expenses(user_id, current_month, current_year)

        # 5. Obtener próximos pagos (reutilizando la lógica de alerts)
        upcoming_payments = get_recurring_payment_reminders(user_id)

        return Response(
            {
                "success": True,
                "data": {
                    "weekly_tip": weekly_tip,
                    "budget": budget_data,
                    "daily_expenses": daily_expenses,
                    "upcoming_payments": upcoming_payments
                }
            },
            status=status.HTTP_200_OK
        )

    def _get_weekly_tip(self, user_id: int):
        """Obtiene el tip semanal personalizado basado en IA"""
        try:
            tip_message = get_or_generate_weekly_tip(user_id)
            return {
                "title": "Consejo de la semana",
                "message": tip_message
            }
        except Exception as e:
            # Fallback en caso de error
            print(f"Error obteniendo weekly tip: {str(e)}")
            return {
                "title": "Consejo de ahorro",
                "message": "Revisa tus gastos semanalmente para mantener el control de tu presupuesto."
            }

    def _get_budget_data(self, user_id, month, year):
        """Calcular presupuesto del mes actual"""

        # Total gastado este mes
        total_gastado = Expense.objects.filter(
            user_id=user_id,
            date__month=month,
            date__year=year
        ).aggregate(total=Sum('total'))['total'] or Decimal('0')

        # Presupuesto restante (balance actual)
        try:
            balance = Balance.objects.get(user_id=user_id)
            presupuesto_restante = float(balance.current_amount)
        except Balance.DoesNotExist:
            presupuesto_restante = 0

        # Total del presupuesto
        total_presupuesto = float(total_gastado) + presupuesto_restante

        # Calcular porcentaje gastado
        if total_presupuesto > 0:
            porcentaje_gastado = round((float(total_gastado) / total_presupuesto) * 100, 1)
        else:
            porcentaje_gastado = 0

        return {
            "month": month,  # Número del mes (1-12)
            "year": year,
            "total": round(total_presupuesto, 2),
            "spent": round(float(total_gastado), 2),
            "remaining": round(presupuesto_restante, 2),
            "percentage": porcentaje_gastado
        }

    def _get_daily_expenses(self, user_id, month, year):
        """
        Obtener gastos agrupados por día del mes actual
        Retorna array con {day, amount} para cada día que tuvo gastos
        """

        # Obtener todos los gastos del mes
        expenses = Expense.objects.filter(
            user_id=user_id,
            date__month=month,
            date__year=year
        ).values('date').annotate(
            total_amount=Sum('total')
        ).order_by('date')

        # Formatear datos para el gráfico
        daily_data = []
        for expense in expenses:
            daily_data.append({
                "day": expense['date'].day,
                "amount": round(float(expense['total_amount']), 2)
            })

        return daily_data

class MonthlyExpensesPredictionView(APIView):
    permission_classes = [AllowAny]
    """
    Obtiene gastos mensuales reales y predicciones para el resto del año
    """

    def get(self, request):
        user_id = request.query_params.get('user_id')
        year = request.query_params.get('year', None)

        if not user_id:
            return Response(
                {"error": "user_id es requerido"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            if not year:
                from datetime import datetime
                year = datetime.now().year
            else:
                year = int(year)

        except ValueError as e:
            return Response(
                {"error": f"Parámetros inválidos: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        from datetime import datetime
        current_date = datetime.now()
        current_month = current_date.month if current_date.year == year else 12

        # Obtener gastos reales de cada mes
        monthly_expenses = []
        for month in range(1, 13):
            month_total = Expense.objects.filter(
                user_id=user_id,
                date__month=month,
                date__year=year
            ).aggregate(total=Sum('total'))['total']

            monthly_expenses.append({
                "month": month,
                "real": float(month_total) if month_total else None
            })

        # Calcular predicción para meses futuros
        months_with_data = [m for m in monthly_expenses if m["real"] is not None and m["real"] > 0]

        if len(months_with_data) >= 3:
            last_3_months = months_with_data[-3:]
            avg_expense = sum(m["real"] for m in last_3_months) / 3
        elif len(months_with_data) > 0:
            avg_expense = sum(m["real"] for m in months_with_data) / len(months_with_data)
        else:
            avg_expense = 0

        # Obtener el valor real del mes actual para usar como punto de partida
        current_month_value = monthly_expenses[current_month - 1]["real"] if current_month <= 12 else None

        predictions = []
        for i, month_data in enumerate(monthly_expenses):
            month = month_data["month"]

            if month < current_month:
                # Meses pasados - no hay predicción
                predictions.append(None)
            elif month == current_month:
                # Mes actual - usar valor real para conectar las líneas
                predictions.append(current_month_value)
            else:
                # Meses futuros - calcular predicción
                months_ahead = month - current_month
                # Usar el valor del mes actual como base
                base_value = current_month_value if current_month_value else avg_expense
                predicted_value = base_value * (1.02 ** months_ahead)
                predictions.append(round(predicted_value, 2))

        # Formatear respuesta
        result = []
        for i, month_data in enumerate(monthly_expenses):
            result.append({
                "month": month_data["month"],
                "real": month_data["real"],
                "prediction": predictions[i]
            })

        return Response(
            {
                "success": True,
                "data": result,
                "year": year,
                "current_month": current_month
            },
            status=status.HTTP_200_OK
        )

class ExpensesByCategoryView(APIView):
    permission_classes = [AllowAny]
    """
    Obtiene el total gastado por cada categoría en un mes específico
    """

    def get(self, request):
        user_id = request.query_params.get('user_id')
        month = request.query_params.get('month')  # Formato: número del 1-12
        year = request.query_params.get('year', None)  # Opcional, por defecto año actual

        if not user_id or not month:
            return Response(
                {"error": "user_id y month son requeridos"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            month = int(month)
            if month < 1 or month > 12:
                raise ValueError("Mes inválido")

            # Si no se proporciona year, usar el año actual
            if not year:
                from datetime import datetime
                year = datetime.now().year
            else:
                year = int(year)

        except ValueError as e:
            return Response(
                {"error": f"Parámetros inválidos: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Filtrar expenses por user, mes y año
        expenses = Expense.objects.filter(
            user_id=user_id,
            date__month=month,
            date__year=year
        )

        # Agrupar por categoría y sumar totales
        category_totals = expenses.values('category').annotate(
            total=Sum('total')
        )

        # Crear un diccionario con todas las categorías inicializadas en 0
        all_categories = {cat.value: 0 for cat in Category}

        # Actualizar con los totales reales
        for item in category_totals:
            all_categories[item['category']] = float(item['total'])

        # Formatear respuesta con todas las categorías
        result = [
            {
                "category": category,
                "total": total
            }
            for category, total in all_categories.items()
        ]

        return Response(
            {
                "success": True,
                "data": result,
                "month": month,
                "year": year
            },
            status=status.HTTP_200_OK
        )

class ExpensesByParentCategoryView(APIView):
    permission_classes = [AllowAny]
    """
    Obtiene el total gastado por cada categoría padre en un mes específico
    """

    def get(self, request):
        user_id = request.query_params.get('user_id')
        month = request.query_params.get('month')
        year = request.query_params.get('year', None)

        if not user_id or not month:
            return Response(
                {"error": "user_id y month son requeridos"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            month = int(month)
            if month < 1 or month > 12:
                raise ValueError("Mes inválido")

            if not year:
                from datetime import datetime
                year = datetime.now().year
            else:
                year = int(year)

        except ValueError as e:
            return Response(
                {"error": f"Parámetros inválidos: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Filtrar expenses por user, mes y año
        expenses = Expense.objects.filter(
            user_id=user_id,
            date__month=month,
            date__year=year
        )

        # Inicializar totales por categoría padre en 0
        parent_totals = {parent.value: 0 for parent in CategoryParent}

        # Sumar gastos por cada expense
        for expense in expenses:
            try:
                # Obtener la categoría del expense
                category = Category(expense.category)
                # Obtener la categoría padre correspondiente
                parent_category = CATEGORY_PARENT_MAP.get(category)

                if parent_category:
                    parent_totals[parent_category.value] += float(expense.total)

            except ValueError:
                # Si la categoría no es válida, continuar
                continue

        # Formatear respuesta
        result = [
            {
                "parent_category": parent_category,
                "total": total
            }
            for parent_category, total in parent_totals.items()
        ]

        return Response(
            {
                "success": True,
                "data": result,
                "month": month,
                "year": year
            },
            status=status.HTTP_200_OK
        )

class SavingsEvolutionView(APIView):
    permission_classes = [AllowAny]
    """
    Obtiene la evolución del ahorro mes a mes
    """

    def get(self, request):
        user_id = request.query_params.get('user_id')
        year = request.query_params.get('year', None)

        if not user_id:
            return Response(
                {"error": "user_id es requerido"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            if not year:
                from datetime import datetime
                year = datetime.now().year
            else:
                year = int(year)

        except ValueError as e:
            return Response(
                {"error": f"Parámetros inválidos: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        from moneymind_apps.balances.models import UserBalanceHistory

        # Obtener el historial de balances del año
        history = UserBalanceHistory.objects.filter(
            user_id=user_id,
            date__year=year
        ).order_by('date')

        # Formatear respuesta
        monthly_savings = []
        previous_amount = None

        for record in history:
            # Calcular el ahorro (diferencia con el mes anterior)
            if previous_amount is not None:
                saving = float(record.amount) - previous_amount
            else:
                saving = 0  # Primer mes no tiene comparación

            monthly_savings.append({
                "month": record.date.month,
                "date": record.date.isoformat(),
                "balance": float(record.amount),
                "saving": saving
            })

            previous_amount = float(record.amount)

        return Response(
            {
                "success": True,
                "data": monthly_savings,
                "year": year
            },
            status=status.HTTP_200_OK
        )

class EssentialVsNonEssentialExpensesView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        user_id = request.query_params.get('user_id')
        year = request.query_params.get('year', None)

        if not user_id:
            return Response(
                {"error": "user_id es requerido"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            if not year:
                from datetime import datetime
                year = datetime.now().year
            else:
                year = int(year)

        except ValueError as e:
            return Response(
                {"error": f"Parámetros inválidos: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        from moneymind_apps.movements.models import ExpenseType, CATEGORY_EXPENSE_TYPE_MAP

        # Resultado por cada mes
        monthly_data = []

        for month in range(1, 13):  # Meses del 1 al 12
            # Filtrar expenses del mes
            expenses = Expense.objects.filter(
                user_id=user_id,
                date__month=month,
                date__year=year
            )

            esencial_total = 0
            no_esencial_total = 0

            # Clasificar y sumar
            for expense in expenses:
                try:
                    category = Category(expense.category)
                    expense_type = CATEGORY_EXPENSE_TYPE_MAP.get(category)

                    if expense_type == ExpenseType.ESENCIAL:
                        esencial_total += float(expense.total)
                    elif expense_type == ExpenseType.NO_ESENCIAL:
                        no_esencial_total += float(expense.total)

                except ValueError:
                    continue

            monthly_data.append({
                "month": month,
                "esencial": esencial_total,
                "no_esencial": no_esencial_total
            })

        return Response(
            {
                "success": True,
                "data": monthly_data,
                "year": year
            },
            status=status.HTTP_200_OK
        )


def get_or_generate_weekly_tip(user_id: int) -> str:
    """
    Obtiene el tip semanal del usuario o genera uno nuevo si no existe o está vencido
    """
    try:
        tip_obj = WeeklyTip.objects.get(user_id=user_id)

        # Si el tip tiene más de 7 días, regenerar
        if tip_obj.should_regenerate():
            new_tip = generate_weekly_tip(user_id)
            tip_obj.tip = new_tip
            tip_obj.created_at = datetime.now()
            tip_obj.save()
            return new_tip

        return tip_obj.tip

    except WeeklyTip.DoesNotExist:
        # No existe, crear uno nuevo
        new_tip = generate_weekly_tip(user_id)
        WeeklyTip.objects.create(user_id=user_id, tip=new_tip)
        return new_tip


class ExportReportView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        # Obtener parámetros
        user_id = request.query_params.get('user_id')
        report_type = request.query_params.get('report_type')
        file_format = request.query_params.get('file_format')

        if not all([user_id, report_type, file_format]):
            return HttpResponse("Parámetros faltantes", status=400)

        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return HttpResponse("Usuario no encontrado", status=404)

        # Determinar rango de fechas según tipo de reporte
        start_date, end_date = self._get_date_range(request, report_type)

        if not start_date or not end_date:
            return HttpResponse("Fechas inválidas", status=400)

        # Obtener datos del reporte
        report_data = self._generate_report_data(user, start_date, end_date)

        # Generar archivo según formato
        if file_format == 'excel':
            return self._generate_excel(report_data, start_date, end_date)
        elif file_format == 'pdf':
            return self._generate_pdf(report_data, start_date, end_date)
        else:
            return HttpResponse("Formato no soportado", status=400)

    def _get_date_range(self, request, report_type):
        """Determina el rango de fechas según el tipo de reporte"""

        if report_type == 'monthly':
            month = int(request.query_params.get('month', timezone.now().month))
            year = int(request.query_params.get('year', timezone.now().year))

            start_date = datetime(year, month, 1).date()
            # Último día del mes
            if month == 12:
                end_date = datetime(year, 12, 31).date()
            else:
                end_date = (datetime(year, month + 1, 1) - timedelta(days=1)).date()

        elif report_type == 'yearly':
            year = int(request.query_params.get('year', timezone.now().year))
            start_date = datetime(year, 1, 1).date()
            end_date = datetime(year, 12, 31).date()

        elif report_type == 'custom':
            start_date_str = request.query_params.get('start_date')
            end_date_str = request.query_params.get('end_date')

            if not start_date_str or not end_date_str:
                return None, None

            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        else:
            return None, None

        return start_date, end_date

    def _generate_report_data(self, user, start_date, end_date):
        """Genera todos los datos necesarios para el reporte"""

        # 1. Obtener ingresos y gastos del período
        incomes = Income.objects.filter(
            user=user,
            date__gte=start_date,
            date__lte=end_date
        ).order_by('date', 'time')

        expenses = Expense.objects.filter(
            user=user,
            date__gte=start_date,
            date__lte=end_date
        ).order_by('date', 'time')

        # 2. Calcular totales
        total_income = incomes.aggregate(total=Sum('total'))['total'] or Decimal('0')
        total_expenses = expenses.aggregate(total=Sum('total'))['total'] or Decimal('0')
        balance = total_income - total_expenses

        # 3. Tasa de ahorro
        savings_rate = 0
        if total_income > 0:
            savings_rate = round((balance / total_income) * 100, 1)

        # 4. Gastos por categoría
        expenses_by_category = expenses.values('category').annotate(
            total=Sum('total'),
            count=Count('id')
        ).order_by('-total')

        # Calcular porcentajes
        category_data = []
        for item in expenses_by_category:
            percentage = 0
            if total_expenses > 0:
                percentage = round((item['total'] / total_expenses) * 100, 1)

            category_data.append({
                'category': item['category'],
                'total': float(item['total']),
                'percentage': percentage,
                'count': item['count']
            })

        # 5. Movimientos detallados
        movements = []

        # Agregar ingresos
        for income in incomes:
            movements.append({
                'date': income.date,
                'time': income.time,
                'type': 'Ingreso',
                'description': income.title,
                'amount': float(income.total),
                'is_income': True
            })

        # Agregar gastos
        for expense in expenses:
            movements.append({
                'date': expense.date,
                'time': expense.time,
                'type': 'Gasto',
                'category': expense.category,
                'description': expense.place,
                'amount': float(expense.total),
                'is_income': False
            })

        # Ordenar por fecha y hora
        movements.sort(key=lambda x: (x['date'], x['time']))

        # 6. Estadísticas adicionales
        total_days = (end_date - start_date).days + 1
        avg_daily_expense = float(total_expenses) / total_days if total_days > 0 else 0

        # Día con más gastos
        day_expenses = expenses.values('date').annotate(
            total=Sum('total')
        ).order_by('-total').first()

        max_expense_day = None
        max_expense_amount = 0
        if day_expenses:
            max_expense_day = day_expenses['date']
            max_expense_amount = float(day_expenses['total'])

        # Categoría más frecuente
        most_frequent_category = None
        most_frequent_count = 0
        if category_data:
            most_frequent = max(category_data, key=lambda x: x['count'])
            most_frequent_category = most_frequent['category']
            most_frequent_count = most_frequent['count']

        total_transactions = incomes.count() + expenses.count()

        return {
            'user': {
                'name': user.get_full_name() or user.username,
                'email': user.email
            },
            'period': {
                'start': start_date,
                'end': end_date,
                'label': self._get_period_label(start_date, end_date)
            },
            'summary': {
                'total_income': float(total_income),
                'total_expenses': float(total_expenses),
                'balance': float(balance),
                'savings_rate': savings_rate
            },
            'expenses_by_category': category_data,
            'movements': movements,
            'statistics': {
                'avg_daily_expense': round(avg_daily_expense, 2),
                'max_expense_day': max_expense_day,
                'max_expense_amount': max_expense_amount,
                'most_frequent_category': most_frequent_category,
                'most_frequent_count': most_frequent_count,
                'total_transactions': total_transactions
            }
        }

    def _get_period_label(self, start_date, end_date):
        """Genera etiqueta del período"""
        months_es = [
            'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
            'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
        ]

        if start_date.year == end_date.year and start_date.month == end_date.month:
            # Mismo mes
            return f"{months_es[start_date.month - 1]} {start_date.year}"
        elif start_date.month == 1 and end_date.month == 12 and start_date.year == end_date.year:
            # Año completo
            return f"Año {start_date.year}"
        else:
            # Rango personalizado
            return f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"

    def _generate_excel(self, data, start_date, end_date):
        """Genera archivo Excel con el reporte financiero"""

        # Crear workbook
        wb = Workbook()
        wb.remove(wb.active)  # Eliminar hoja por defecto

        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14)
        center_align = Alignment(horizontal='center', vertical='center')
        currency_format = '"S/"#,##0.00'

        # ============ HOJA 1: RESUMEN ============
        ws_summary = wb.create_sheet("Resumen")

        # Título
        ws_summary['A1'] = "REPORTE FINANCIERO"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary.merge_cells('A1:D1')

        ws_summary['A2'] = f"Periodo: {data['period']['label']}"
        ws_summary['A2'].font = Font(size=12)
        ws_summary.merge_cells('A2:D2')

        ws_summary['A3'] = f"Usuario: {data['user']['name']}"
        ws_summary.merge_cells('A3:D3')

        # Resumen financiero
        row = 5
        summary_data = [
            ["Concepto", "Monto"],
            ["Ingresos totales", data['summary']['total_income']],
            ["Gastos totales", data['summary']['total_expenses']],
            ["Balance", data['summary']['balance']],
            ["Tasa de ahorro", f"{data['summary']['savings_rate']}%"]
        ]

        for row_data in summary_data:
            ws_summary.append(row_data)

        # Aplicar formato
        for cell in ws_summary[row]:
            cell.fill = header_fill
            cell.font = header_font

        # Formato de moneda para columna B
        for row_idx in range(row + 1, row + 5):
            if row_idx != row + 4:  # No aplicar a tasa de ahorro
                ws_summary[f'B{row_idx}'].number_format = currency_format

        # Ajustar anchos
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 20

        # ============ HOJA 2: GASTOS POR CATEGORÍA ============
        ws_categories = wb.create_sheet("Gastos por Categoría")

        ws_categories['A1'] = "Gastos por Categoría"
        ws_categories['A1'].font = title_font
        ws_categories.merge_cells('A1:D1')

        # Headers
        headers = ["Categoría", "Monto", "% del Total", "Transacciones"]
        ws_categories.append([""])  # Fila vacía
        ws_categories.append(headers)

        header_row = 3
        for cell in ws_categories[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        # Datos
        for cat_data in data['expenses_by_category']:
            ws_categories.append([
                cat_data['category'].replace('_', ' ').title(),
                cat_data['total'],
                f"{cat_data['percentage']}%",
                cat_data['count']
            ])

        # Formato de moneda
        for row_idx in range(4, 4 + len(data['expenses_by_category'])):
            ws_categories[f'B{row_idx}'].number_format = currency_format

        # Ajustar anchos
        ws_categories.column_dimensions['A'].width = 30
        ws_categories.column_dimensions['B'].width = 15
        ws_categories.column_dimensions['C'].width = 15
        ws_categories.column_dimensions['D'].width = 15

        # ============ HOJA 3: MOVIMIENTOS DETALLADOS ============
        ws_movements = wb.create_sheet("Movimientos Detallados")

        ws_movements['A1'] = "Lista Detallada de Movimientos"
        ws_movements['A1'].font = title_font
        ws_movements.merge_cells('A1:F1')

        # Headers
        headers = ["Fecha", "Hora", "Tipo", "Categoría", "Lugar/Desc", "Monto"]
        ws_movements.append([""])
        ws_movements.append(headers)

        header_row = 3
        for cell in ws_movements[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        # Datos
        for movement in data['movements']:
            amount = movement['amount'] if movement['is_income'] else -movement['amount']

            category = movement.get('category', '')  # Evita KeyError
            if category:
                category = category.replace('_', ' ').title()

            description = movement.get('description', '')  # También lo protegemos

            ws_movements.append([
                movement['date'].strftime('%d/%m/%Y'),
                movement['time'],
                movement['type'],
                category,  # ← Nunca causará error
                description,  # ← Protegido también
                amount
            ])

        # Formato de moneda con signo
        for row_idx in range(4, 4 + len(data['movements'])):
            cell = ws_movements[f'F{row_idx}']
            cell.number_format = '"S/"#,##0.00;[Red]-"S/"#,##0.00'

        # Ajustar anchos
        ws_movements.column_dimensions['A'].width = 12
        ws_movements.column_dimensions['B'].width = 8
        ws_movements.column_dimensions['C'].width = 10
        ws_movements.column_dimensions['D'].width = 25
        ws_movements.column_dimensions['E'].width = 30
        ws_movements.column_dimensions['F'].width = 15

        # ============ HOJA 4: ESTADÍSTICAS ============
        ws_stats = wb.create_sheet("Estadísticas")

        ws_stats['A1'] = "Estadísticas Adicionales"
        ws_stats['A1'].font = title_font
        ws_stats.merge_cells('A1:B1')

        stats = data['statistics']

        stats_data = [
            [""],
            ["Métrica", "Valor"],
            ["Gasto promedio diario", f"S/ {stats['avg_daily_expense']:.2f}"],
            ["Día con más gastos",
             f"{stats['max_expense_day'].strftime('%d/%m/%Y')} (S/ {stats['max_expense_amount']:.2f})" if stats[
                 'max_expense_day'] else "N/A"],
            ["Categoría más frecuente",
             f"{stats['most_frequent_category'].replace('_', ' ').title()} ({stats['most_frequent_count']} transacciones)" if
             stats['most_frequent_category'] else "N/A"],
            ["Total de transacciones", f"{stats['total_transactions']} movimientos"]
        ]

        for row_data in stats_data:
            ws_stats.append(row_data)

        # Formato header
        for cell in ws_stats[3]:
            cell.fill = header_fill
            cell.font = header_font

        # Ajustar anchos
        ws_stats.column_dimensions['A'].width = 30
        ws_stats.column_dimensions['B'].width = 40

        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Preparar respuesta
        filename = f"reporte_financiero_{data['period']['start'].strftime('%Y%m%d')}_{data['period']['end'].strftime('%Y%m%d')}.xlsx"

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    def _generate_pdf(self, data, start_date, end_date):
        """Genera archivo PDF con el reporte financiero"""

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter,
                                rightMargin=72, leftMargin=72,
                                topMargin=72, bottomMargin=18)

        # Container para los elementos
        elements = []

        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1033d3'),
            spaceAfter=30,
            alignment=TA_CENTER
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#1033d3'),
            spaceAfter=12,
            spaceBefore=12
        )

        # ============ TÍTULO Y ENCABEZADO ============
        elements.append(Paragraph("REPORTE FINANCIERO", title_style))
        elements.append(Paragraph(f"<b>Periodo:</b> {data['period']['label']}", styles['Normal']))
        elements.append(Paragraph(f"<b>Usuario:</b> {data['user']['name']}", styles['Normal']))
        elements.append(Spacer(1, 20))

        # ============ RESUMEN EJECUTIVO ============
        elements.append(Paragraph("Resumen de Cifras Clave", heading_style))

        summary_data = [
            ['Concepto', 'Monto'],
            ['Ingresos totales', f"S/ {data['summary']['total_income']:,.2f}"],
            ['Gastos totales', f"S/ {data['summary']['total_expenses']:,.2f}"],
            ['Balance (Neto)', f"S/ {data['summary']['balance']:,.2f}"],
            ['Tasa de ahorro', f"{data['summary']['savings_rate']}%"]
        ]

        summary_table = Table(summary_data, colWidths=[3 * inch, 2 * inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1033d3')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # ============ GASTOS POR CATEGORÍA ============
        elements.append(Paragraph("Gastos por Categoría", heading_style))

        if data['expenses_by_category']:
            category_data = [['Categoría', 'Monto', '% del Total']]

            for cat in data['expenses_by_category']:
                category_data.append([
                    cat['category'].replace('_', ' ').title(),
                    f"S/ {cat['total']:,.2f}",
                    f"{cat['percentage']}%"
                ])

            category_table = Table(category_data, colWidths=[3 * inch, 1.5 * inch, 1.5 * inch])
            category_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1033d3')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))

            elements.append(category_table)
        else:
            elements.append(Paragraph("No hay gastos registrados en este período.", styles['Normal']))

        elements.append(PageBreak())

        # ============ MOVIMIENTOS DETALLADOS ============
        elements.append(Paragraph("Lista Detallada de Movimientos", heading_style))

        if data['movements']:
            # Limitar a primeros 30 movimientos para no hacer el PDF demasiado largo
            movements_display = data['movements'][:30]

            movement_data = [['Fecha', 'Hora', 'Tipo', 'Categoría', 'Monto']]

            for mov in movements_display:
                amount_str = f"+S/ {mov['amount']:,.2f}" if mov['is_income'] else f"-S/ {mov['amount']:,.2f}"

                category = mov.get('category', '')  # ← Si no existe, devuelve ''
                if category:
                    category = category.replace('_', ' ').title()[:15]

                movement_data.append([
                    mov['date'].strftime('%d/%m/%Y'),
                    mov['time'],
                    mov['type'],
                    category,  # ← Ya nunca causará KeyError
                    amount_str
                ])

            movement_table = Table(movement_data, colWidths=[1 * inch, 0.7 * inch, 0.8 * inch, 2 * inch, 1.2 * inch])
            movement_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1033d3')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))

            elements.append(movement_table)

            if len(data['movements']) > 30:
                elements.append(Spacer(1, 12))
                elements.append(Paragraph(
                    f"<i>Mostrando 30 de {len(data['movements'])} transacciones. Descargue el reporte en Excel para ver el listado completo.</i>",
                    styles['Normal']
                ))
        else:
            elements.append(Paragraph("No hay movimientos registrados en este período.", styles['Normal']))

        elements.append(Spacer(1, 20))

        # ============ ESTADÍSTICAS ADICIONALES ============
        elements.append(Paragraph("Estadísticas Adicionales", heading_style))

        stats = data['statistics']

        stats_data = [
            ['Métrica', 'Valor'],
            ['Gasto promedio diario', f"S/ {stats['avg_daily_expense']:.2f}"],
            ['Día con más gastos',
             f"{stats['max_expense_day'].strftime('%d/%m/%Y')} (S/ {stats['max_expense_amount']:.2f})" if stats[
                 'max_expense_day'] else "N/A"],
            ['Categoría más frecuente',
             f"{stats['most_frequent_category'].replace('_', ' ').title()} ({stats['most_frequent_count']} transacciones)" if
             stats['most_frequent_category'] else "N/A"],
            ['Total de transacciones', f"{stats['total_transactions']} movimientos"]
        ]

        stats_table = Table(stats_data, colWidths=[2.5 * inch, 3.5 * inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1033d3')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        elements.append(stats_table)

        # Construir PDF
        doc.build(elements)

        # Preparar respuesta
        buffer.seek(0)
        filename = f"reporte_financiero_{data['period']['start'].strftime('%Y%m%d')}_{data['period']['end'].strftime('%Y%m%d')}.pdf"

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response